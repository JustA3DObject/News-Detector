import requests
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import re
import sys
import os

page_link = []
articles_list = []
link_list = []

news_type = input(
    "Select if you want to fetch for true news or false/parody/satire news. (Enter 1 for true snd 0 false): ")
try:
    int(news_type)

except ValueError:
    print("Invalid input! Please enter integer type values.")
    sys.exit(0)

news_type = int(news_type)

if news_type == 1:
    file_name = 'True News.xlsx'
elif news_type == 0:
    file_name = 'False News.xlsx'
else:
    print("Invalid input! Please input either 1 or 0.")
    sys.exit(0)

website = input(
    "Choose website to fetch news from. (Enter 1 for Hindustan Times, enter 2 for NDTV or enter 3 for The Indian Express): ")

try:
    int(website)

except ValueError:
    print("Invalid input! Please enter integer type values.")
    sys.exit(0)

website = int(website)

if website == 1:
    website = "hindustantimes"
elif website == 2:
    website = "ndtv"
elif website == 3:
    website = "theindianexpress"
else:
    print("Invalid input! Please enter values in range (1,2 or 3).")
    sys.exit(0)

category = input(
    f'Please enter the topic keyword corresponding to the website {website} for fetching articles: ')

number_of_pages = input(
    "Enter number of pages to fetch articles from (Max limit is 50 for Hindustan Times, 14 for NDTV and 100 for The Indian Express): ")

try:
    int(number_of_pages)

except ValueError:
    print("Invalid input! Please enter integer type values.")
    sys.exit(0)

number_of_pages = int(number_of_pages)

if website == "hindustantimes" and number_of_pages > 50:
    print("Number of pages are out of range. Value set to maximum (50)")
    number_of_pages = 50
elif website == "ndtv" and number_of_pages > 14:
    print("Number of pages are out of range. Value set to maximum (14)")
    number_of_pages = 14
elif website == "theindianexpress" and number_of_pages > 100:
    print("Number of pages are out of range. Value set to maximum (100)")
    number_of_pages = 100

headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.114 Safari/537.36",
           "X-Amzn-Trace-Id": "Root=1-62d8036d-2b173c1f2e4e7a416cc9e554", "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
           "Accept-Encoding": "gzip, deflate, br", "Accept-Language": "en-GB", }

fetch_values_dictionary = {
    "hindustantimes": {
        "domain_link": "https://www.hindustantimes.com/",
        "link": f'https://www.hindustantimes.com/{category}/page-',
        "range_start": 1,
        "range_end": 1 + number_of_pages,
        "articles_find_element": {
            "element": "section",
            "element_class": "listingPage"
        },
        "articles_findall_element": {
            "element": "div",
            "element_class": "cartHolder",
        },
        "link_find_element": {
            "element": "h3",
            "element_class": "hdg3",
        },
        "content_find_element": {
            "element": "div",
            "element_class": "detailPage",
        },
        "title_find_element": {
            "element": "h1",
            "element_class": "hdg1",
        },
        "datetime_find_element": {
            "element": "div",
            "element_class": "dateTime",
        },
    },
    "ndtv": {
        "domain_link": "https://www.ndtv.com/",
        "link": f'https://www.ndtv.com/{category}/page-',
        "range_start": 1,
        "range_end": 1 + number_of_pages,
        "articles_find_element": {
            "element": "div",
            "element_class": "lisingNews"
        },
        "articles_findall_element": {
            "element": "div",
            "element_class": "news_Itm",
        },
        "link_find_element": {
            "element": "h2",
            "element_class": "newsHdng",
        },
        "content_find_element": {
            "element": "section",
            "element_class": "col-900",
        },
        "title_find_element": {
            "element": "h1",
            "element_class": "sp-ttl",
        },
        "datetime_find_element": {
            "element": "span",
            "element_itemprop": "dateModified",
        },
    },
    "theindianexpress": {
        "domain_link": "https://indianexpress.com/",
        "link": f'https://indianexpress.com/section/{category}/page/',
        "range_start": 2,
        "range_end": 2 + number_of_pages,
        "articles_find_element": {
            "element": "div",
            "element_class": "nation"
        },
        "articles_findall_element": {
            "element": "div",
            "element_class": "articles",
        },
        "link_find_element": {
            "element": "div",
            "element_class": "snaps",
        },
        "content_find_element": {
            "element": "div",
            "element_class": "container native_story",
        },
        "title_find_element": {
            "element": "h1",
            "element_itemprop": "headline",
        },
        "datetime_find_element": {
            "element": "span",
            "element_itemprop": "dateModified",
        },
    },
}


if os.path.exists(file_name):
    excel = openpyxl.load_workbook(file_name)
    sheet = excel.active
else:
    excel = openpyxl.Workbook()
    sheet = excel.active
    sheet.title = 'News'
    sheet.append(['Title', 'Text', 'Subject', 'Date'])


def fetch_news(fetch_values_dictionary, category, website):

    if len(re.findall('news', category)) == 0:
        category = f'{category} News'
    else:
        category.replace('-news', ' News')
    category = category.capitalize()

    def article_fetcher():

        print("Fetching Articles...")

        for i in range(fetch_values_dictionary[website].get("range_start"),
                       fetch_values_dictionary[website].get("range_end")):
            page_link = (
                f'{fetch_values_dictionary[website].get("link")}{i}')

            i = i+1
            try:
                page = requests.get(page_link, headers=headers)
                page.raise_for_status()

                if page.status_code == 200:
                    soup1 = BeautifulSoup(page.text, "html.parser")
                    soup2 = BeautifulSoup(soup1.prettify(), "html.parser")

                articles = soup2.find(fetch_values_dictionary[website]["articles_find_element"].get("element"),
                                      class_=fetch_values_dictionary[website]["articles_find_element"].get("element_class")).find_all(
                    fetch_values_dictionary[website]["articles_findall_element"].get(
                        "element"),
                    class_=fetch_values_dictionary[website]["articles_findall_element"].get("element_class"))

                articles_list.extend(articles)

            except Exception as e:
                print(e)

    article_fetcher()

    def link_list_maker():

        print("Getting Links...")

        for article in articles_list:

            link = article.find(fetch_values_dictionary[website]["link_find_element"].get("element"),
                                class_=fetch_values_dictionary[website]["link_find_element"].get("element_class"))

            if link is not None:
                link = link.find('a').get('href')

                if (len(re.findall(fetch_values_dictionary[website].get("domain_link"), link)) == 0):
                    link = f'{fetch_values_dictionary[website].get("domain_link")}{link}'
                link_list.append(link)

    link_list_maker()

    def content_fetcher():

        print("Fetching Content...")

        for i in range(len(link_list)):
            try:
                page = requests.get(link_list[i], headers=headers)
                page.raise_for_status()

                if page.status_code == 200:
                    soup1 = BeautifulSoup(page.text, "html.parser")
                    soup2 = BeautifulSoup(soup1.prettify(), "html.parser")

                if fetch_values_dictionary[website]["content_find_element"].get("element_class") is not None:
                    content = soup2.find(
                        fetch_values_dictionary[website]["content_find_element"].get(
                            "element"),
                        class_=fetch_values_dictionary[website]["content_find_element"].get("element_class"))
                elif fetch_values_dictionary[website]["content_find_element"].get("element_id") is not None:
                    content = soup2.find(
                        fetch_values_dictionary[website]["content_find_element"].get(
                            "element"),
                        id=fetch_values_dictionary[website]["content_find_element"].get("element_id"))
                elif fetch_values_dictionary[website]["content_find_element"].get("element_itemprop") is not None:
                    content = soup2.find(
                        fetch_values_dictionary[website]["content_find_element"].get(
                            "element"),
                        itemprop=fetch_values_dictionary[website]["content_find_element"].get("element_itemprop"))

                try:
                    if fetch_values_dictionary[website]["title_find_element"].get("element_class") is not None:
                        title = content.find(
                            fetch_values_dictionary[website]["title_find_element"].get(
                                "element"),
                            class_=fetch_values_dictionary[website]["title_find_element"].get("element_class")).get_text().strip()
                    elif fetch_values_dictionary[website]["title_find_element"].get("element_id") is not None:
                        title = content.find(
                            fetch_values_dictionary[website]["title_find_element"].get(
                                "element"),
                            id=fetch_values_dictionary[website]["title_find_element"].get("element_id")).get_text().strip()
                    elif fetch_values_dictionary[website]["title_find_element"].get("element_itemprop") is not None:
                        title = content.find(
                            fetch_values_dictionary[website]["title_find_element"].get(
                                "element"),
                            itemprop=fetch_values_dictionary[website]["title_find_element"].get("element_itemprop")).get_text().strip()

                    title = ' '.join(title.split())

                except AttributeError:
                    title = "Null"

                try:
                    if fetch_values_dictionary[website]["datetime_find_element"].get("element_class") is not None:
                        date_time = content.find(
                            fetch_values_dictionary[website]["datetime_find_element"].get(
                                "element"), class_=fetch_values_dictionary[website]["datetime_find_element"].get("element_class")).get_text().strip()
                    elif fetch_values_dictionary[website]["datetime_find_element"].get("element_id") is not None:
                        date_time = content.find(
                            fetch_values_dictionary[website]["datetime_find_element"].get(
                                "element"), id=fetch_values_dictionary[website]["datetime_find_element"].get("element_id")).get_text().strip()
                    elif fetch_values_dictionary[website]["datetime_find_element"].get("element_itemprop") is not None:
                        date_time = content.find(
                            fetch_values_dictionary[website]["datetime_find_element"].get(
                                "element"), itemprop=fetch_values_dictionary[website]["datetime_find_element"].get("element_itemprop")).get_text().strip()

                    date_time = date_time.replace('Updated: ', '')

                except AttributeError:
                    date_time = "Null"

                try:
                    body = [x.get_text() for x in content.find_all('p')]
                    body = ' '.join(body)
                    body = ' '.join(body.split())
                    body = body.replace(' Watch Live News: Follow Us:', '')
                    body = body.replace(' ...view detail', '')
                except AttributeError:
                    body = "Null"

                if (title and body and date_time) != "Null":
                    sheet.append([title, body, category, date_time])
                # print(title, date_time, category, body)

            except Exception as e:
                print(e)

    content_fetcher()


fetch_news(fetch_values_dictionary, category, website)

print("Saving Excel File...")
excel.save(file_name)
